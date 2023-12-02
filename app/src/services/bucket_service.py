from aiogram.types import InlineKeyboardButton
from aiogram.utils.keyboard import InlineKeyboardBuilder
from openpyxl import Workbook, load_workbook
from sqlalchemy import insert

from core.const import base_buttons
from db.shemas import Order, OrderItems
from services.base_service import BaseService

class BucketService(BaseService):

    def show_bucket_itmes(self, bucket: dict) -> str:
        sorted_bucket = sorted(bucket.values(), key=lambda x: x.get('name'))
        res = ''
        for i, ele in enumerate(sorted_bucket, 1):
            cur_str = f"{i}. Наименование: {ele.get('name')}.\n"
            if ele.get('description'):
                cur_str += f"Описание: {ele.get('description')}.\n"
            cur_str += f"Количество: {ele.get('quantity')}.\n\n"
            res += cur_str
        return res

    def show_for_delete_items(self, bucket: dict) -> InlineKeyboardBuilder:

        sorted_bucket_list = sorted(bucket.items(), key=lambda x: x[0])
        lst = [(f"❌ {ele[1].get('name')}", ele[0]) for ele in sorted_bucket_list]
        keybord = self.create_keybord(lst)
        keybord.row(*[InlineKeyboardButton(text='Вернуться в корзину',
                                           callback_data='bucket')])
        return keybord

    async def create_order(self, user_id: int, address: str,
                           bucket: dict) -> None:
        order = Order(client_id=user_id, address=address)
        self.db.add(order)
        await self.db.flush()
        await self.db.refresh(order)
        items = bucket.values()
        stmt = [{'order_id': order.id,
                 'item_id': ele.get('id'),
                 'quantity': int(ele.get('quantity'))} for ele in items]
        values_for_excel = [[str(order.id), user_id, ele.get('id'), ele.get('name'),
                            ele.get('quantity'), address] for ele in items]
        await self.db.execute(insert(OrderItems), stmt)
        filename = "orders/orders.xlsx"
        try:
            wb = load_workbook(filename)
            ws = wb.worksheets[0]
        except FileNotFoundError:
            headers_row = ['order_id', 'user_id', 'item_id', 'item name',
                           'quantity', 'address']
            wb = Workbook()
            ws = wb.active
            ws.append(headers_row)
        finally:
            for ele in values_for_excel:
                ws.append(ele)
            wb.save(filename)
        await self.db.commit()